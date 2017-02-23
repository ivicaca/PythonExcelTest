from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
import glob

dest_filename = 'empty_book.xlsx'
wbEmpty = Workbook()
wsEmpty=wbEmpty.active
i=0
#read excels
for name in glob.glob("test*.xlsx"):
    print(name)
    wb1=load_workbook(name)
    sheet1=wb1.get_sheet_by_name("prvi")
    i += 1
    col_0=sheet1.columns[0]
    for idx, cell in enumerate(col_0, 1):
        wsEmpty.cell(row=idx,column=i).value = cell.value
wbEmpty.save("trt.xlsx")






